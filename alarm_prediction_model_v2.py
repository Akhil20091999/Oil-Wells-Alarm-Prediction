# ─────────────────────────────────────────────────────────────────────────────
#  ALARM PREDICTION MODEL  –  Oil Well Analytics
#  Predicts which alarms will fire in the NEXT 1 HOUR
#  Uses all sensor readings + alarm limits as inputs
#  Output: console list + Excel file of predicted alarms
# ─────────────────────────────────────────────────────────────────────────────

import pandas as pd
import numpy as np
import re
import warnings
from datetime import timedelta
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, accuracy_score
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────
EXCEL_FILE    = 'E:\\AKHIL\\Data Analytics\\Alarm Prediction\\OilWell_Plant_Data.xlsx'
SHEET_SENSOR  = 'Sensor Time Series'
PREDICT_HOURS = 1                  # predict alarms in next N hours
WINDOW_STEPS  = 4                  # look-back window = 4 × 15 min = 1 hour
MIN_PROB      = 0.40               # minimum probability to call as "alarm likely"
OUTPUT_EXCEL  = 'Alarm_Predictions_Next1Hour.xlsx'

# ──────────────────────────────────────────────────────────────────────────────
# STEP 1: LOAD SENSOR DATA + ALARM LIMITS
# ──────────────────────────────────────────────────────────────────────────────
print("=" * 65)
print("  OIL WELL  –  1-HOUR ALARM PREDICTION MODEL")
print("=" * 65)
print("\n[1/6] Loading sensor data and alarm limits...")

raw = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_SENSOR, header=None)

# Row 1 = tag names, Row 2 = limits text, Row 3+ = data
tag_names = [str(t).split('\n')[0].strip() for t in raw.iloc[1, 1:].tolist()]
limit_row = raw.iloc[2, 1:].tolist()

LIMITS = {
    # ── Pressure sensors (psi) ─────────────────────────────────────────
    'PT-1001': {'LL': 200, 'LO': 300, 'HI': 3500, 'HH': 4000},  # Wellhead Tubing Pressure
    'PT-1002': {'LL': 150, 'LO': 250, 'HI': 2800, 'HH': 3200},  # Wellhead Casing Pressure
    'PT-1003': {'LL': 50, 'LO': 100, 'HI': 1200, 'HH': 1500},  # Flowline Pressure
    'PT-2001': {'LL': 20, 'LO': 40, 'HI': 700, 'HH': 850},  # Separator Inlet Pressure
    'PT-3002': {'LL': 100, 'LO': 200, 'HI': 1800, 'HH': 2200},  # Compressor Discharge Pressure

    # ── Temperature sensors (°F) ───────────────────────────────────────
    'TT-1001': {'LL': 60, 'LO': 80, 'HI': 280, 'HH': 320},  # Wellhead Flowing Temperature
    'TT-3001': {'LL': 50, 'LO': 80, 'HI': 380, 'HH': 430},  # Compressor Discharge Temp
    'TT-3002': {'LL': 40, 'LO': 60, 'HI': 200, 'HH': 240},  # Pump Bearing Temperature
    'TT-3003': {'LL': 50, 'LO': 70, 'HI': 260, 'HH': 300},  # Motor Winding Temperature

    # ── Flow sensors ───────────────────────────────────────────────────
    'FT-1001': {'LL': 100, 'LO': 200, 'HI': 8000, 'HH': 10000},  # Gross Liquid Flow (bbl/d)
    'FT-1002': {'LL': 50, 'LO': 100, 'HI': 5000, 'HH': 6500},  # Oil Flow Rate (bbl/d)
    'FT-1003': {'LL': 0.05, 'LO': 0.1, 'HI': 15, 'HH': 20},  # Gas Flow Rate (MMscfd)
    'FT-1004': {'LL': 0, 'LO': 10, 'HI': 4000, 'HH': 5000},  # Water Flow Rate (bbl/d)

    # ── Level sensors (%) ──────────────────────────────────────────────
    'LT-2001': {'LL': 5, 'LO': 15, 'HI': 80, 'HH': 90},  # Separator Oil Level
    'LT-2002': {'LL': 5, 'LO': 10, 'HI': 85, 'HH': 95},  # Separator Water Level
    'LT-3001': {'LL': 3, 'LO': 10, 'HI': 88, 'HH': 95},  # Produced Water Tank Level

    # ── Speed & Vibration ─────────────────────────────────────────────
    'ST-3001': {'LL': 100, 'LO': 300, 'HI': 3400, 'HH': 3600},  # ESP Motor Speed (rpm)
    'VT-3001': {'LL': 0, 'LO': 0, 'HI': 12, 'HH': 18},  # Compressor Vibration X (mm/s)
    'VT-3003': {'LL': 0, 'LO': 0, 'HI': 8, 'HH': 12},  # Pump Vibration (mm/s)

    # ── Gas Detection ─────────────────────────────────────────────────
    'GD-4001': {'LL': 0, 'LO': 0, 'HI': 20, 'HH': 40},  # LEL Gas Detector (% LEL)
    'GD-4002': {'LL': 0, 'LO': 0, 'HI': 5, 'HH': 10},  # H2S Concentration (ppm)

    # ── Electrical ────────────────────────────────────────────────────
    'ET-5001': {'LL': 5, 'LO': 10, 'HI': 180, 'HH': 220},  # ESP Motor Current (A)
    'ET-5002': {'LL': 350, 'LO': 380, 'HI': 480, 'HH': 510},  # ESP Motor Voltage (V)
    'ET-5003': {'LL': 32, 'LO': 50, 'HI': 140, 'HH': 160},  # VFD Drive Temperature (°F)

    # ── Safety ────────────────────────────────────────────────────────
    'PS-6001': {'LL': 0, 'LO': 0, 'HI': 3800, 'HH': 4200},  # HIPPS Pressure (psi)

    # ── Quality ───────────────────────────────────────────────────────
    'AT-7001': {'LL': 0, 'LO': 0, 'HI': 60, 'HH': 75},  # Water Cut (%)

    # ── Integrity ────────────────────────────────────────────────────
    'CT-8001': {'LL': 0, 'LO': 0, 'HI': 8, 'HH': 15},  # Corrosion Rate (mpy)
    'CT-8003': {'LL': 0, 'LO': 0, 'HI': 50, 'HH': 100},  # Sand Production Rate (lb/d)
}

# Tag descriptions (from tag reference)
DESCRIPTIONS = {
    'PT-1001': 'Wellhead Tubing Pressure',
    'PT-1002': 'Wellhead Casing Pressure',
    'PT-1003': 'Flowline Pressure',
    'PT-2001': 'Separator Inlet Pressure',
    'PT-3002': 'Compressor Discharge Pressure',
    'TT-1001': 'Wellhead Flowing Temperature',
    'TT-3001': 'Compressor Discharge Temp',
    'TT-3002': 'Pump Bearing Temperature',
    'TT-3003': 'Motor Winding Temperature',
    'FT-1001': 'Gross Liquid Flow',
    'FT-1002': 'Oil Flow Rate',
    'FT-1003': 'Gas Flow Rate',
    'FT-1004': 'Water Flow Rate',
    'LT-2001': 'Separator Oil Level',
    'LT-2002': 'Separator Water Level',
    'LT-3001': 'Produced Water Tank Level',
    'ST-3001': 'ESP Motor Speed',
    'VT-3001': 'Compressor Vibration X',
    'VT-3003': 'Pump Vibration',
    'GD-4001': 'LEL Gas Detector',
    'GD-4002': 'H2S Concentration',
    'ET-5001': 'ESP Motor Current',
    'ET-5002': 'ESP Motor Voltage',
    'ET-5003': 'VFD Drive Temperature',
    'PS-6001': 'HIPPS Pressure',
    'AT-7001': 'Water Cut',
    'CT-8001': 'Corrosion Rate',
    'CT-8003': 'Sand Production Rate',
}

UNITS = {
    'PT-1001':'psi','PT-1002':'psi','PT-1003':'psi','PT-2001':'psi','PT-3002':'psi',
    'TT-1001':'°F','TT-3001':'°F','TT-3002':'°F','TT-3003':'°F',
    'FT-1001':'bbl/d','FT-1002':'bbl/d','FT-1003':'MMscfd','FT-1004':'bbl/d',
    'LT-2001':'%','LT-2002':'%','LT-3001':'%',
    'ST-3001':'rpm','VT-3001':'mm/s','VT-3003':'mm/s',
    'GD-4001':'% LEL','GD-4002':'ppm',
    'ET-5001':'A','ET-5002':'V','ET-5003':'°F',
    'PS-6001':'psi','AT-7001':'%','CT-8001':'mpy','CT-8003':'lb/d',
}

PRIORITIES = {
    'PT-1001':'Critical','PT-1002':'High','PT-1003':'High','PT-2001':'Critical',
    'PT-3002':'Critical','TT-1001':'High','TT-3001':'Critical','TT-3002':'High',
    'TT-3003':'Critical','FT-1001':'High','FT-1002':'High','FT-1003':'High',
    'FT-1004':'Medium','LT-2001':'High','LT-2002':'High','LT-3001':'Critical',
    'ST-3001':'Critical','VT-3001':'Critical','VT-3003':'High',
    'GD-4001':'Critical','GD-4002':'Critical','ET-5001':'Critical',
    'ET-5002':'High','ET-5003':'High','PS-6001':'Critical',
    'AT-7001':'High','CT-8001':'High','CT-8003':'Critical',
}

# Load data rows (skip rows 0-2 which are headers)
df = raw.iloc[3:].copy()
df.columns = ['Timestamp'] + tag_names
df['Timestamp'] = pd.to_datetime(df['Timestamp'])
df = df.reset_index(drop=True)
for tag in tag_names:
    df[tag] = pd.to_numeric(df[tag], errors='coerce')

print(f"      Loaded {len(df):,} rows  ×  {len(tag_names)} sensors")
print(f"      Period: {df['Timestamp'].min().date()} → {df['Timestamp'].max().date()}")

# ──────────────────────────────────────────────────────────────────────────────
# STEP 2: ENGINEER FEATURES
# ──────────────────────────────────────────────────────────────────────────────
print("\n[2/6] Engineering features for each sensor tag...")

# For every sensor, compute these features at each timestep:
#   - current value
#   - normalised distance to HI limit  (value / HI)
#   - normalised distance to LL limit  (LL / value, clamped)
#   - rate of change (difference from 1 step ago)
#   - rolling mean over last WINDOW_STEPS
#   - rolling max over last WINDOW_STEPS
#   - already in alarm (1/0)
#   - hour of day, day of week

feature_cols = []

for tag in tag_names:
    lim = LIMITS[tag]
    hi  = lim['HI'] if lim['HI'] > 0 else 1.0
    hh  = lim['HH'] if lim['HH'] > 0 else 1.0
    ll  = lim['LL']
    lo  = lim['LO']
    v   = df[tag]

    # Normalised proximity to HI/HH (1.0 = at limit, >1 = over limit)
    df[f'{tag}_prox_hi']  = (v / hi).clip(0, 3)
    df[f'{tag}_prox_hh']  = (v / hh).clip(0, 3)
    feature_cols += [f'{tag}_prox_hi', f'{tag}_prox_hh']

    # Proximity to LL/LO (only meaningful if limits > 0)
    if lo > 0:
        df[f'{tag}_prox_lo'] = (lo / v.replace(0, np.nan)).clip(0, 3).fillna(3)
        df[f'{tag}_prox_ll'] = (ll / v.replace(0, np.nan)).clip(0, 3).fillna(3) if ll > 0 else 0
        feature_cols += [f'{tag}_prox_lo', f'{tag}_prox_ll']

    # Rate of change (delta from previous step)
    df[f'{tag}_roc'] = v.diff(1).fillna(0)
    feature_cols.append(f'{tag}_roc')

    # Rolling mean and max over look-back window
    df[f'{tag}_roll_mean'] = v.rolling(WINDOW_STEPS, min_periods=1).mean()
    df[f'{tag}_roll_max']  = v.rolling(WINDOW_STEPS, min_periods=1).max()
    df[f'{tag}_roll_min']  = v.rolling(WINDOW_STEPS, min_periods=1).min()
    feature_cols += [f'{tag}_roll_mean', f'{tag}_roll_max', f'{tag}_roll_min']

    # Already in alarm right now?
    in_alarm = (
        (v >= hh) | (v >= hi) |
        ((lo > 0) & (v <= lo)) |
        ((ll > 0) & (v <= ll))
    ).astype(int)
    df[f'{tag}_in_alarm'] = in_alarm
    feature_cols.append(f'{tag}_in_alarm')

# Time features
df['hour']        = df['Timestamp'].dt.hour
df['day_of_week'] = df['Timestamp'].dt.dayofweek
feature_cols += ['hour', 'day_of_week']

print(f"      Built {len(feature_cols)} features across {len(tag_names)} tags")

# ──────────────────────────────────────────────────────────────────────────────
# STEP 3: BUILD TARGET LABELS
# ──────────────────────────────────────────────────────────────────────────────
print("\n[3/6] Building prediction targets (will alarm in next 1 hour?)...")

PREDICT_STEPS = PREDICT_HOURS * 4   # 4 steps per hour at 15-min freq

def will_alarm_in_window(series, hi, hh, lo, ll, steps):
    """For each row, check if any future value within 'steps' crosses a limit."""
    target = np.zeros(len(series), dtype=int)
    arr = series.values
    for i in range(len(arr) - steps):
        future_slice = arr[i+1 : i+1+steps]
        if (np.any(future_slice >= hi) or np.any(future_slice >= hh) or
           (lo > 0 and np.any(future_slice <= lo)) or
           (ll > 0 and np.any(future_slice <= ll))):
            target[i] = 1
    return target

targets = {}
for tag in tag_names:
    lim = LIMITS[tag]
    targets[tag] = will_alarm_in_window(
        df[tag], lim['HI'], lim['HH'], lim['LO'], lim['LL'], PREDICT_STEPS
    )
    df[f'target_{tag}'] = targets[tag]
    pct = targets[tag].mean() * 100
    print(f"      {tag:10s}  alarm-in-next-1h: {targets[tag].sum():4d} / {len(df):,}  ({pct:.1f}%)")

# ──────────────────────────────────────────────────────────────────────────────
# STEP 4: TRAIN ONE RANDOM FOREST PER TAG
# ──────────────────────────────────────────────────────────────────────────────
print("\n[4/6] Training Random Forest model per sensor tag...")

# Drop rows with NaNs in features
feat_df = df[feature_cols].fillna(0)

models    = {}
tag_scores = {}

for tag in tag_names:
    y = df[f'target_{tag}']

    # Skip if almost no positive examples (nothing to learn)
    if y.sum() < 10:
        print(f"      {tag:10s}  SKIP (too few alarm events)")
        continue

    X_train, X_test, y_train, y_test = train_test_split(
        feat_df, y, test_size=0.2, random_state=42, shuffle=False
    )

    clf = RandomForestClassifier(
        n_estimators=80,
        max_depth=8,
        min_samples_leaf=5,
        class_weight='balanced',   # handles imbalanced alarm data
        random_state=42,
        n_jobs=-1
    )
    clf.fit(X_train, y_train)
    y_pred = clf.predict(X_test)
    acc    = accuracy_score(y_test, y_pred)

    models[tag]     = clf
    tag_scores[tag] = acc
    print(f"      {tag:10s}  accuracy: {acc*100:.1f}%  "
          f"(+events: {y.sum():4d})")

print(f"\n      Models trained: {len(models)} / {len(tag_names)} tags")

# ──────────────────────────────────────────────────────────────────────────────
# STEP 5: PREDICT ON MOST RECENT DATA WINDOW
# ──────────────────────────────────────────────────────────────────────────────
print("\n[5/6] Running predictions on the latest sensor readings...")

# Use the last row as "current" live data
# In production you would plug in real-time sensor values here
last_row = feat_df.iloc[[-1]]    # shape (1, n_features) – most recent reading

predictions = []

for tag, clf in models.items():
    lim  = LIMITS[tag]
    prob = clf.predict_proba(last_row)[0]
    # prob[1] = probability of alarm occurring in next 1 hour
    prob_alarm = prob[1] if len(prob) > 1 else 0.0
    current_val = float(df[tag].iloc[-1])

    # Determine likely alarm type based on proximity
    hi, hh = lim['HI'], lim['HH']
    lo, ll = lim['LO'], lim['LL']

    dist_hh = (current_val / hh) if hh > 0 else 0
    dist_hi = (current_val / hi) if hi > 0 else 0
    dist_lo = (lo / current_val) if (lo > 0 and current_val > 0) else 0
    dist_ll = (ll / current_val) if (ll > 0 and current_val > 0) else 0

    # Pick most likely alarm direction
    distances = {
        'HH': dist_hh,
        'HI': dist_hi,
        'LO': dist_lo,
        'LL': dist_ll,
    }
    likely_type = max(distances, key=distances.get)

    # Trend: rate of change over last 4 steps
    roc = float(df[f'{tag}_roc'].iloc[-1])
    trend = '↑ Rising' if roc > 0.5 else ('↓ Falling' if roc < -0.5 else '→ Stable')

    # Proximity to nearest limit (as %)
    if likely_type in ('HH','HI'):
        limit_val  = hh if likely_type == 'HH' else hi
        prox_pct   = round((current_val / limit_val) * 100, 1) if limit_val else 0
        margin     = round(limit_val - current_val, 2)
    else:
        limit_val  = ll if likely_type == 'LL' else lo
        prox_pct   = round((limit_val / current_val) * 100, 1) if current_val else 0
        margin     = round(current_val - limit_val, 2)

    predictions.append({
        'Tag ID':           tag,
        'Description':      DESCRIPTIONS.get(tag, tag),
        'Category':         next((c for t,d,a,u,ll2,lo2,hi2,hh2,p,c,nm,nx
                                  in [] if t==tag), 'Sensor'),
        'Priority':         PRIORITIES.get(tag, 'Medium'),
        'Unit':             UNITS.get(tag, ''),
        'Current Value':    round(current_val, 3),
        'Alarm Probability %': round(prob_alarm * 100, 1),
        'Likely Alarm Type':   likely_type,
        'Nearest Limit':    limit_val,
        'Margin to Limit':  margin,
        '% of Limit':       prox_pct,
        'Trend':            trend,
        'Model Accuracy %': round(tag_scores.get(tag, 0) * 100, 1),
        'LL': ll, 'LO': lo, 'HI': hi, 'HH': hh,
    })

# Sort by alarm probability descending
predictions.sort(key=lambda x: x['Alarm Probability %'], reverse=True)

# ──────────────────────────────────────────────────────────────────────────────
# STEP 6: CONSOLE OUTPUT
# ──────────────────────────────────────────────────────────────────────────────
print("\n[6/6] Prediction results...")

HIGH_RISK   = [p for p in predictions if p['Alarm Probability %'] >= 70]
MEDIUM_RISK = [p for p in predictions if 40 <= p['Alarm Probability %'] < 70]
LOW_RISK    = [p for p in predictions if p['Alarm Probability %'] < 40]

def print_section(title, items, emoji):
    if not items: return
    print(f"\n  {emoji}  {title}  ({len(items)} tags)")
    print("  " + "─" * 90)
    print(f"  {'Tag':<10} {'Description':<32} {'Priority':<10} "
          f"{'Current':>10} {'Unit':<8} {'Prob%':>6} {'Type':<5} "
          f"{'Margin':>10} {'Trend':<12}")
    print("  " + "─" * 90)
    for p in items:
        print(f"  {p['Tag ID']:<10} {p['Description']:<32} "
              f"{p['Priority']:<10} {p['Current Value']:>10.3f} "
              f"{p['Unit']:<8} {p['Alarm Probability %']:>6.1f} "
              f"{p['Likely Alarm Type']:<5} {p['Margin to Limit']:>10.2f} "
              f"{p['Trend']:<12}")

print("\n" + "═"*65)
print("  PREDICTED ALARMS  –  NEXT 1 HOUR")
print(f"  Based on latest reading: {df['Timestamp'].iloc[-1]}")
print("═"*65)

print_section("HIGH RISK  (≥70% probability)",  HIGH_RISK,   "🔴")
print_section("MEDIUM RISK  (40–69%)",           MEDIUM_RISK, "🟠")
print_section("LOW RISK  (<40%)",                LOW_RISK,    "🟢")

print(f"\n  Total tags assessed : {len(predictions)}")
print(f"  High risk alarms    : {len(HIGH_RISK)}")
print(f"  Medium risk alarms  : {len(MEDIUM_RISK)}")
print(f"  Low risk alarms     : {len(LOW_RISK)}")

# ──────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ──────────────────────────────────────────────────────────────────────────────
print(f"\n  Saving Excel report → {OUTPUT_EXCEL} ...")

wb  = Workbook()
thin = Side(style='thin', color='CCCCCC')
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, text, bg='1F3864', fg='FFFFFF', sz=10, wrap=False):
    cell.value     = text
    cell.font      = Font(name='Arial', size=sz, bold=True, color=fg)
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center',
                               wrap_text=wrap)
    cell.border    = bdr

def dat(cell, val, fg='333333', bg='FFFFFF', bold=False, fmt=None, wrap=False):
    cell.value     = val
    cell.font      = Font(name='Arial', size=9, bold=bold, color=fg)
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center',
                               wrap_text=wrap)
    cell.border    = bdr
    if fmt:
        cell.number_format = fmt

# ── Risk colours ──────────────────────────────────────────────────────────────
PRI_BG  = {'Critical':'FFCCCC','High':'FFE8CC','Medium':'FFFACC','Low':'E8FFE8'}
PRI_FG  = {'Critical':'CC0000','High':'C55A11','Medium':'7D6608','Low':'1A5276'}
TYPE_FG = {'HH':'CC0000','HI':'C55A11','LO':'1565C0','LL':'1A237E'}
TYPE_BG = {'HH':'FFD0D0','HI':'FFE8D0','LO':'D0E8FF','LL':'D8D0FF'}

def prob_color(p):
    if p >= 70: return ('CC0000','FFCCCC')    # red
    if p >= 40: return ('C55A11','FFE8CC')    # orange
    return ('1A5276','E8F5FB')                # blue

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1: Predicted Alarm List
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Predicted Alarms – Next 1 Hour'

# Title block
ws1.merge_cells('A1:O1')
title_cell = ws1.cell(1,1)
title_cell.value = (f'OIL WELL ALARM PREDICTIONS — NEXT 1 HOUR  |  '
                    f'Based on: {df["Timestamp"].iloc[-1]}')
title_cell.font      = Font(name='Arial', size=13, bold=True, color='FFFFFF')
title_cell.fill      = PatternFill('solid', start_color='1A2744')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 28

# KPI row
kpi_labels = [
    ('Total Tags Assessed', len(predictions)),
    ('🔴 HIGH RISK (≥70%)', len(HIGH_RISK)),
    ('🟠 MEDIUM RISK (40-69%)', len(MEDIUM_RISK)),
    ('🟢 LOW RISK (<40%)', len(LOW_RISK)),
    ('Critical Priority Alarms', sum(1 for p in HIGH_RISK+MEDIUM_RISK
                                      if p['Priority']=='Critical')),
]
kpi_cols = [1,4,7,10,13]
for (lbl, val), c in zip(kpi_labels, kpi_cols):
    ws1.merge_cells(start_row=2, start_column=c, end_row=2, end_column=c+2)
    ws1.merge_cells(start_row=3, start_column=c, end_row=3, end_column=c+2)
    cell_l = ws1.cell(2, c); cell_v = ws1.cell(3, c)
    cell_l.value = lbl
    cell_l.font  = Font(name='Arial', size=9, bold=True, color='666666')
    cell_l.alignment = Alignment(horizontal='center')
    cell_v.value = val
    cell_v.font  = Font(name='Arial', size=20, bold=True, color='1A2744')
    cell_v.alignment = Alignment(horizontal='center')
ws1.row_dimensions[2].height = 18
ws1.row_dimensions[3].height = 30

# Column headers (row 4)
COL_HDRS = [
    'Rank','Tag ID','Description','Priority','Unit',
    'Current Value','Alarm Prob %','Likely Alarm Type',
    'Nearest Limit','Margin to Limit','% of Limit',
    'Trend','LL','LO','HI','HH','Model Accuracy %'
]
COL_WIDTHS = [6,10,30,10,9,13,13,15,13,15,12,12,8,8,8,8,16]

for ci, h in enumerate(COL_HDRS, 1):
    hdr(ws1.cell(4, ci), h, wrap=True)
ws1.row_dimensions[4].height = 32

# Data rows
for rank, p in enumerate(predictions, 1):
    row = 4 + rank
    prob = p['Alarm Probability %']
    pri  = p['Priority']
    at   = p['Likely Alarm Type']
    pfg, pbg = prob_color(prob)

    row_bg = ('FFF5F5' if prob >= 70 else
              'FFF9F0' if prob >= 40 else 'F5F9FF')

    vals = [
        rank,
        p['Tag ID'],
        p['Description'],
        p['Priority'],
        p['Unit'],
        p['Current Value'],
        prob,
        at,
        p['Nearest Limit'],
        p['Margin to Limit'],
        p['% of Limit'],
        p['Trend'],
        p['LL'], p['LO'], p['HI'], p['HH'],
        p['Model Accuracy %'],
    ]
    for ci, val in enumerate(vals, 1):
        cell = ws1.cell(row, ci)
        col_name = COL_HDRS[ci-1]

        if col_name == 'Priority':
            dat(cell, val, fg=PRI_FG.get(pri,'333333'),
                bg=PRI_BG.get(pri,'FFFFFF'), bold=True)
        elif col_name == 'Alarm Prob %':
            dat(cell, val, fg=pfg, bg=pbg, bold=True)
            cell.number_format = '0.0"%"'
        elif col_name == 'Likely Alarm Type':
            dat(cell, val, fg=TYPE_FG.get(at,'333333'),
                bg=TYPE_BG.get(at,'FFFFFF'), bold=True)
        elif col_name == 'Trend':
            trend_fg = ('CC0000' if '↑' in str(val) and at in ('HH','HI')
                        else '1565C0' if '↓' in str(val) and at in ('LL','LO')
                        else '333333')
            dat(cell, val, fg=trend_fg, bg=row_bg)
        elif col_name in ('LL','LO','HI','HH'):
            dat(cell, val, fg='555555', bg='F8F8F8')
        else:
            dat(cell, val, bg=row_bg)

for ci, w in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w

ws1.freeze_panes = 'A5'
ws1.auto_filter.ref = f'A4:{get_column_letter(len(COL_HDRS))}4'

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2: High-Risk Detail (≥70%)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('High Risk Detail')

ws2.merge_cells('A1:I1')
c = ws2.cell(1,1)
c.value = '🔴  HIGH RISK ALARMS — Probability ≥ 70%  |  Requires Immediate Attention'
c.font  = Font(name='Arial', size=12, bold=True, color='FFFFFF')
c.fill  = PatternFill('solid', start_color='CC0000')
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 26

detail_hdrs = ['Tag ID','Description','Priority','Current Value','Unit',
               'Alarm Prob %','Likely Alarm Type','Nearest Limit',
               'Margin','LL','LO','HI','HH','Trend','Recommended Action']
detail_widths = [10,30,10,13,8,13,15,13,12,8,8,8,8,12,45]

for ci, h in enumerate(detail_hdrs, 1):
    hdr(ws2.cell(2, ci), h, bg='7B0000', wrap=True)
ws2.row_dimensions[2].height = 30

ACTIONS = {
    'HH': 'IMMEDIATE: Reduce {tag} – value approaching critical HH limit. Check upstream cause.',
    'HI': 'ALERT: Monitor {tag} closely – trending toward HI. Consider corrective action.',
    'LO': 'ALERT: {tag} falling – check supply/feed source. Approaching LO limit.',
    'LL': 'IMMEDIATE: {tag} critically low. Check for line blockage or equipment failure.',
}

for ri, p in enumerate(HIGH_RISK, 3):
    at  = p['Likely Alarm Type']
    pri = p['Priority']
    action = ACTIONS.get(at,'').format(tag=p['Description'])
    row_vals = [
        p['Tag ID'], p['Description'], pri,
        p['Current Value'], p['Unit'],
        p['Alarm Probability %'], at,
        p['Nearest Limit'], p['Margin to Limit'],
        p['LL'], p['LO'], p['HI'], p['HH'],
        p['Trend'], action
    ]
    for ci, val in enumerate(row_vals, 1):
        cell = ws2.cell(ri, ci)
        col  = detail_hdrs[ci-1]
        if col == 'Priority':
            dat(cell, val, fg=PRI_FG.get(pri,'333333'),
                bg=PRI_BG.get(pri,'FFFFFF'), bold=True)
        elif col == 'Alarm Prob %':
            fg2, bg2 = prob_color(val)
            dat(cell, val, fg=fg2, bg=bg2, bold=True)
        elif col == 'Likely Alarm Type':
            dat(cell, val, fg=TYPE_FG.get(at,'333333'),
                bg=TYPE_BG.get(at,'FFFFFF'), bold=True)
        elif col == 'Recommended Action':
            dat(cell, val, fg='333333', bg='FFF9F0', wrap=True)
            ws2.row_dimensions[ri].height = 32
        else:
            dat(cell, val, bg='FFF5F5')

for ci, w in enumerate(detail_widths, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.freeze_panes = 'A3'

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3: Model Performance
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Model Performance')

ws3.merge_cells('A1:F1')
c = ws3.cell(1,1)
c.value = 'RANDOM FOREST MODEL ACCURACY PER TAG'
c.font  = Font(name='Arial', size=12, bold=True, color='FFFFFF')
c.fill  = PatternFill('solid', start_color='1F3864')
c.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 26

perf_hdrs = ['Tag ID','Description','Priority','Training Samples',
             'Model Accuracy %','Alarm Events in Data']
for ci, h in enumerate(perf_hdrs, 1):
    hdr(ws3.cell(2, ci), h)
ws3.row_dimensions[2].height = 28

for ri, tag in enumerate(tag_names, 3):
    if tag not in models: continue
    acc    = round(tag_scores[tag]*100, 1)
    n_pos  = int(targets[tag].sum())
    pri    = PRIORITIES.get(tag,'Medium')
    acc_fg = ('1A5276' if acc >= 85 else
              'C55A11' if acc >= 70 else 'CC0000')
    row_vals = [tag, DESCRIPTIONS.get(tag,tag), pri,
                int(len(df)*0.8), acc, n_pos]
    for ci, val in enumerate(row_vals, 1):
        cell = ws3.cell(ri, ci)
        if perf_hdrs[ci-1] == 'Priority':
            dat(cell, val, fg=PRI_FG.get(pri,'333333'),
                bg=PRI_BG.get(pri,'FFFFFF'), bold=True)
        elif perf_hdrs[ci-1] == 'Model Accuracy %':
            dat(cell, val, fg=acc_fg, bold=True,
                bg=('E8F5E9' if acc>=85 else
                    'FFF8E1' if acc>=70 else 'FFEBEE'))
        else:
            dat(cell, val, bg='F8F8F8' if ri%2==0 else 'FFFFFF')

for ci, w in enumerate([10,30,10,16,16,20], 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.freeze_panes = 'A3'

wb.save(OUTPUT_EXCEL)
print(f"  ✅  Saved → {OUTPUT_EXCEL}")
print("\n" + "═"*65)
print("  DONE.  Open Alarm_Predictions_Next1Hour.xlsx to review.")
print("═"*65)
